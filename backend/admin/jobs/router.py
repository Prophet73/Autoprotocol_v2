"""
Admin Jobs Router - управление очередью задач.

Active jobs (pending/processing) come from Redis for real-time status.
Historical jobs (completed/failed) come from PostgreSQL for permanent storage.
"""
import os
import logging
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func, desc
from sqlalchemy.orm import selectinload

from backend.core.auth import require_admin
from backend.core.storage import get_job_store
from backend.shared.database import get_db_readonly, get_db
from backend.shared.models import User, TranscriptionJob

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/jobs", tags=["Админ - Задачи"])


@router.get("", summary="Список всех задач")
async def list_all_jobs(
    limit: int = Query(default=50, ge=1, le=1000),
    page: int = Query(default=1, ge=1),
    status: Optional[str] = Query(default=None, description="Filter by status"),
    domain: Optional[str] = Query(default=None, description="Filter by domain"),
    user_email: Optional[str] = Query(default=None, description="Filter by user email"),
    _: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db_readonly),
):
    """
    Список задач из PostgreSQL (вечное хранение) + Redis (активные).

    Активные задачи (pending/processing) обогащаются прогрессом из Redis.
    """
    # Build query from PostgreSQL
    query = (
        select(TranscriptionJob)
        .options(selectinload(TranscriptionJob.user))
    )

    # Apply filters
    if status:
        query = query.where(TranscriptionJob.status == status)
    if domain:
        query = query.where(TranscriptionJob.domain == domain)
    if user_email:
        query = query.join(User, TranscriptionJob.user_id == User.id).where(
            User.email.ilike(f"%{user_email}%")
        )

    # Count total
    count_q = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_q)
    total = total_result.scalar() or 0

    # Paginate
    offset = (page - 1) * limit
    query = query.order_by(desc(TranscriptionJob.created_at)).offset(offset).limit(limit)
    result = await db.execute(query)
    db_jobs = result.scalars().all()

    # Get Redis store for active job enrichment
    store = get_job_store()

    jobs_data = []
    for job in db_jobs:
        uploader_email = job.user.email if job.user else None

        # For active jobs, enrich with Redis progress
        progress_percent = 100 if job.status == "completed" else 0
        current_stage = None
        message = None
        error = job.error_message

        effective_status = job.status

        if job.status in ("pending", "processing"):
            redis_job = store.get(job.job_id)
            if redis_job:
                progress_percent = redis_job.progress_percent or 0
                current_stage = redis_job.current_stage
                message = redis_job.message
                error = redis_job.error
            elif job.created_at and (
                datetime.now(timezone.utc) - job.created_at.replace(tzinfo=timezone.utc)
            ).total_seconds() > 3600:
                # Stale job: pending/processing for >1h with no Redis entry
                effective_status = "failed"
                error = error or "Stale: stuck in queue, no active worker"

        # Build artifacts from DB artifacts JSON + check disk
        artifacts = {}
        if job.artifacts and isinstance(job.artifacts, dict):
            data_dir = Path(os.getenv("DATA_DIR", "/data"))
            output_dir = data_dir / "output" / job.job_id
            for artifact_type, present in job.artifacts.items():
                if present and output_dir.exists():
                    # Find actual file on disk
                    for f in output_dir.iterdir():
                        if artifact_type in f.name.lower() or _matches_artifact(f.name, artifact_type):
                            artifacts[artifact_type] = f.name
                            break

        # Also check Redis output_files for active/recent jobs
        if not artifacts:
            redis_job = store.get(job.job_id)
            if redis_job and redis_job.output_files:
                artifacts = {
                    ftype: Path(fpath).name
                    for ftype, fpath in redis_job.output_files.items()
                }

        jobs_data.append({
            "job_id": job.job_id,
            "status": effective_status,
            "domain": job.domain,
            "meeting_type": job.meeting_type,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "source_file": job.source_filename,
            "progress_percent": progress_percent,
            "current_stage": current_stage,
            "message": message,
            "uploader_email": uploader_email,
            "error": error,
            "artifacts": artifacts,
            "processing_time_seconds": job.processing_time_seconds,
        })

    return {
        "jobs": jobs_data,
        "total": total,
        "page": page,
        "page_size": limit,
    }


def _matches_artifact(filename: str, artifact_type: str) -> bool:
    """Check if filename matches an artifact type."""
    name_lower = filename.lower()
    mapping = {
        "transcript": ["transcript", "стенограмма"],
        "tasks": ["tasks", "задач", ".xlsx"],
        "report": ["report", "отчёт", "отчет"],
        "risk_brief": ["risk", "риск"],
        "summary": ["summary", "конспект"],
    }
    patterns = mapping.get(artifact_type, [artifact_type])
    return any(p in name_lower for p in patterns)


@router.get("/{job_id}/report", summary="JSON-отчёт задачи")
async def get_job_report(
    job_id: str,
    _: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db_readonly),
):
    """
    Получить данные отчёта для просмотра в админке.

    Construction: basic_report + risk_brief из construction_reports (JSON).
    Другие домены: парсим DOCX/XLSX файлы с диска.
    Стенограмма: .txt файл с диска.
    """
    # 1. Check job exists
    job_result = await db.execute(
        select(TranscriptionJob).where(TranscriptionJob.job_id == job_id)
    )
    job = job_result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    report_data: dict = {
        "job_id": job_id,
        "domain": job.domain,
        "meeting_type": job.meeting_type,
        "basic_report": None,
        "risk_brief": None,
        "transcript_text": None,
        "report_text": None,
        "tasks_data": None,
    }

    # 2. Try construction_reports table (has structured JSON)
    try:
        from backend.domains.construction.models import ConstructionReportDB
        cr_result = await db.execute(
            select(ConstructionReportDB).where(ConstructionReportDB.job_id == job_id)
        )
        cr = cr_result.scalar_one_or_none()
        if cr:
            report_data["basic_report"] = cr.basic_report_json
            report_data["risk_brief"] = cr.risk_brief_json
    except Exception as e:
        logger.debug(f"No construction report for {job_id}: {e}")

    # 3. Read files from disk
    data_dir = Path(os.getenv("DATA_DIR", "/data"))
    output_dir = data_dir / "output" / job_id
    if output_dir.exists():
        for f in output_dir.iterdir():
            try:
                # Transcript .txt
                if f.suffix == ".txt" and f.stat().st_size < 500_000:
                    report_data["transcript_text"] = f.read_text(encoding="utf-8")

                # DOCX report — extract text for non-construction domains
                elif f.suffix == ".docx" and not report_data["basic_report"]:
                    if "report" in f.name.lower() or "отчет" in f.name.lower():
                        report_data["report_text"] = _extract_docx_text(f)

                # XLSX tasks — extract rows for non-construction domains
                elif f.suffix == ".xlsx" and not report_data["basic_report"]:
                    if "report" in f.name.lower() or "tasks" in f.name.lower():
                        report_data["tasks_data"] = _extract_xlsx_data(f)

            except Exception as e:
                logger.debug(f"Failed to read {f.name}: {e}")

    return report_data


def _extract_docx_text(filepath: Path) -> str:
    """Extract plain text from a DOCX file."""
    try:
        from docx import Document
        doc = Document(str(filepath))
        paragraphs = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                # Bold paragraphs are likely headings
                if para.runs and para.runs[0].bold:
                    paragraphs.append(f"\n## {text}")
                else:
                    paragraphs.append(text)
        return "\n".join(paragraphs)
    except Exception as e:
        logger.warning(f"Failed to parse DOCX {filepath}: {e}")
        return ""


def _extract_xlsx_data(filepath: Path) -> list[dict]:
    """Extract rows from an XLSX file as list of dicts."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h or "").strip() for h in rows[0]]
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i] and val is not None:
                        row_dict[headers[i]] = str(val)
                if row_dict:
                    result.append(row_dict)
        wb.close()
        return result if result else None
    except Exception as e:
        logger.warning(f"Failed to parse XLSX {filepath}: {e}")
        return None


@router.delete("/{job_id}", summary="Отменить задачу")
async def cancel_job(
    job_id: str,
    _: User = Depends(require_admin),
):
    """
    Принудительно отменить задачу (для админов).
    """
    from backend.core.transcription.models import JobStatus

    store = get_job_store()
    job = store.get(job_id)

    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status not in [JobStatus.PENDING, JobStatus.PROCESSING]:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot cancel job with status: {job.status}"
        )

    # Mark as failed with cancellation message
    store.fail(job_id, "Cancelled by admin")

    # Try to revoke Celery task
    try:
        from backend.tasks.celery_app import celery_app

        celery_app.control.revoke(job_id, terminate=True, signal='SIGTERM')
        logger.info(f"Job {job_id} cancelled by admin and task revoked")
    except Exception as e:
        logger.warning(f"Could not revoke Celery task: {e}")

    return {"success": True, "message": "Job cancelled by admin"}


@router.delete("/{job_id}/purge", summary="Полное удаление задачи")
async def purge_job(
    job_id: str,
    _: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Полное удаление задачи: Redis, PostgreSQL (TranscriptionJob + domain reports),
    файлы (uploads + output). Только для admin/superuser.
    """
    deleted = {"redis": False, "db_job": False, "db_report": False, "uploads": False, "output": False}

    # 1. Delete from Redis
    store = get_job_store()
    deleted["redis"] = store.delete(job_id)

    # 2. Delete from TranscriptionJob table
    result = await db.execute(
        delete(TranscriptionJob).where(TranscriptionJob.job_id == job_id)
    )
    deleted["db_job"] = result.rowcount > 0

    # 3. Delete from construction_reports (if exists)
    try:
        from backend.domains.construction.models import ConstructionReportDB
        result = await db.execute(
            delete(ConstructionReportDB).where(ConstructionReportDB.job_id == job_id)
        )
        deleted["db_report"] = result.rowcount > 0
    except Exception as e:
        logger.warning(f"Could not delete construction report for {job_id}: {e}")

    await db.commit()

    # 4. Delete files from disk
    data_dir = Path(os.getenv("DATA_DIR", "/data"))
    for subdir, key in [("uploads", "uploads"), ("output", "output")]:
        job_dir = data_dir / subdir / job_id
        if job_dir.exists():
            try:
                shutil.rmtree(job_dir)
                deleted[key] = True
                logger.info(f"Purged {subdir} dir: {job_dir}")
            except Exception as e:
                logger.error(f"Failed to delete {job_dir}: {e}")

    if not any(deleted.values()):
        raise HTTPException(status_code=404, detail="Job not found anywhere")

    logger.info(f"Job {job_id} purged by admin: {deleted}")
    return {"success": True, "deleted": deleted}
