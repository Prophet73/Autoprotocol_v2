"""
Excel export for statistics.

Generates multi-sheet XLSX workbook:
1. Сводка — KPI metrics summary
2. По доменам — Domain breakdown
3. Журнал ошибок — Failed jobs log
4. Сырые данные — All jobs raw data
"""
import logging
from io import BytesIO
from datetime import datetime, date, timedelta, timezone
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from backend.shared.models import TranscriptionJob, User
from .schemas import StatsFilters, GeminiPricing

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="0D2C54", end_color="0D2C54", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
BOLD_FONT = Font(bold=True)

from backend.domains.registry import get_display_names as _get_display_names

# Используем функцию для получения актуальных названий из единого реестра
DOMAIN_NAMES = _get_display_names()


def _auto_width(sheet):
    """Auto-adjust column widths to content."""
    for col in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 50)


def _style_headers(sheet):
    """Apply header styling to first row."""
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


async def generate_stats_xlsx(
    db: AsyncSession,
    filters: StatsFilters,
) -> Tuple[BytesIO, str]:
    """Generate Excel workbook with statistics.

    Returns (BytesIO buffer, filename).
    """
    # Resolve date range
    end_date = filters.date_to or date.today()
    start_date = filters.date_from or (end_date - timedelta(days=30))

    # Build conditions
    conditions = [
        TranscriptionJob.created_at >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc),
        TranscriptionJob.created_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc),
    ]
    if filters.domain:
        conditions.append(TranscriptionJob.domain == filters.domain)

    where = and_(*conditions)

    # Fetch all jobs for the period
    result = await db.execute(
        select(TranscriptionJob)
        .where(where)
        .order_by(TranscriptionJob.created_at.desc())
    )
    all_jobs = result.scalars().all()

    # Prefetch users
    user_ids = {j.user_id for j in all_jobs if j.user_id}
    users_map = {}
    if user_ids:
        users_result = await db.execute(
            select(User).where(User.id.in_(user_ids))
        )
        for u in users_result.scalars().all():
            users_map[u.id] = u

    wb = Workbook()

    # ===== Sheet 1: Сводка =====
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.append([f"Сводный отчёт: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"])
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:C1")
    ws1.append([])

    total_jobs = len(all_jobs)
    done_jobs = [j for j in all_jobs if j.status == "completed"]
    errors = [j for j in all_jobs if j.status == "failed"]
    active_users = len({j.user_id for j in all_jobs if j.user_id})

    total_processing = sum(j.processing_time_seconds or 0 for j in done_jobs)
    avg_processing = (total_processing / len(done_jobs)) if done_jobs else 0

    total_input = sum(j.input_tokens or 0 for j in all_jobs)
    total_output = sum(j.output_tokens or 0 for j in all_jobs)
    flash_in_total = sum(getattr(j, 'flash_input_tokens', 0) or 0 for j in all_jobs)
    flash_out_total = sum(getattr(j, 'flash_output_tokens', 0) or 0 for j in all_jobs)
    pro_in_total = sum(getattr(j, 'pro_input_tokens', 0) or 0 for j in all_jobs)
    pro_out_total = sum(getattr(j, 'pro_output_tokens', 0) or 0 for j in all_jobs)
    total_cost = GeminiPricing.calculate_cost_precise(flash_in_total, flash_out_total, pro_in_total, pro_out_total)
    if total_cost == 0 and (total_input > 0 or total_output > 0):
        total_cost = GeminiPricing.calculate_cost(total_input, total_output)

    error_stages = [j.error_stage for j in errors if j.error_stage]
    if error_stages:
        from collections import Counter
        top_error_stage = Counter(error_stages).most_common(1)[0][0]
    else:
        top_error_stage = "—"

    metrics = [
        ("Метрика", "Значение", "Комментарий"),
        ("Активные пользователи", active_users, "С хотя бы 1 задачей"),
        ("Общее количество обработок", total_jobs, f"Среднее на пользователя: {(total_jobs / active_users) if active_users else 0:.1f}"),
        ("Успешных обработок", f"{(len(done_jobs) / total_jobs * 100) if total_jobs else 0:.1f}%", f"Всего: {len(done_jobs)} шт."),
        ("Процент ошибок", f"{(len(errors) / total_jobs * 100) if total_jobs else 0:.1f}%", f"Топ-стадия: {top_error_stage}"),
        ("Среднее время обработки", f"{avg_processing:.1f} сек", f"Всего: {total_processing / 3600:.2f} ч"),
        ("Всего токенов AI (вход)", f"{total_input:,}".replace(",", " "), ""),
        ("Всего токенов AI (выход)", f"{total_output:,}".replace(",", " "), ""),
        ("Стоимость AI", f"${total_cost:.4f}", f"Среднее: ${(total_cost / len(done_jobs)) if done_jobs else 0:.4f}/задача"),
    ]
    for row_data in metrics:
        ws1.append(row_data)
    for cell in ws1[3]:
        cell.font = BOLD_FONT

    # ===== Sheet 2: По доменам =====
    ws2 = wb.create_sheet("По доменам")
    ws2.append(["Домен", "Кол-во", "Успешных", "Ошибок", "Успешность (%)", "Сред. время (сек)", "Токены вход", "Токены выход", "Стоимость ($)"])

    from collections import defaultdict
    domain_stats = defaultdict(lambda: {"count": 0, "done": 0, "failed": 0, "times": [], "input": 0, "output": 0})
    for j in all_jobs:
        d = j.domain or "unknown"
        domain_stats[d]["count"] += 1
        if j.status == "completed":
            domain_stats[d]["done"] += 1
            if j.processing_time_seconds:
                domain_stats[d]["times"].append(j.processing_time_seconds)
        if j.status == "failed":
            domain_stats[d]["failed"] += 1
        domain_stats[d]["input"] += j.input_tokens or 0
        domain_stats[d]["output"] += j.output_tokens or 0

    for dom, stats in sorted(domain_stats.items(), key=lambda x: x[1]["count"], reverse=True):
        avg_t = sum(stats["times"]) / len(stats["times"]) if stats["times"] else 0
        success = (stats["done"] / stats["count"] * 100) if stats["count"] else 0
        cost = GeminiPricing.calculate_cost(stats["input"], stats["output"])
        ws2.append([
            DOMAIN_NAMES.get(dom, dom),
            stats["count"],
            stats["done"],
            stats["failed"],
            f"{success:.1f}%",
            f"{avg_t:.1f}",
            stats["input"],
            stats["output"],
            f"${cost:.4f}",
        ])

    # ===== Sheet 3: Журнал ошибок =====
    ws3 = wb.create_sheet("Журнал ошибок")
    ws3.append(["Время", "Домен", "Пользователь", "Имя файла", "Стадия", "Сообщение об ошибке"])
    for job in sorted(errors, key=lambda j: j.created_at, reverse=True):
        user = users_map.get(job.user_id)
        user_name = user.full_name or user.email if user else "Аноним"
        ws3.append([
            job.created_at.strftime("%Y-%m-%d %H:%M") if job.created_at else "",
            DOMAIN_NAMES.get(job.domain, job.domain or ""),
            user_name,
            job.source_filename or "",
            job.error_stage or "",
            (job.error_message or "")[:500],
        ])

    # ===== Sheet 4: Сырые данные =====
    ws4 = wb.create_sheet("Сырые данные")
    ws4.append([
        "ID", "Job ID", "Время", "Домен", "Тип встречи", "Пользователь",
        "Файл", "Размер (МБ)", "Длит. аудио (сек)", "Длит. аудио (ч)", "Статус",
        "Время обр. (сек)", "Сегменты", "Спикеры",
        "Стадия ошибки", "Ошибка",
        "Токены вход", "Токены выход",
    ])
    for job in all_jobs:
        user = users_map.get(job.user_id)
        user_name = user.full_name or user.email if user else "Аноним"
        size_mb = round(job.source_size_bytes / (1024 * 1024), 2) if job.source_size_bytes else None
        audio_sec = job.audio_duration_seconds
        audio_h = round(audio_sec / 3600, 3) if audio_sec else None
        ws4.append([
            job.id,
            job.job_id,
            job.created_at.strftime("%Y-%m-%d %H:%M") if job.created_at else "",
            DOMAIN_NAMES.get(job.domain, job.domain or ""),
            job.meeting_type or "",
            user_name,
            job.source_filename or "",
            size_mb,
            round(audio_sec, 1) if audio_sec else None,
            audio_h,
            job.status,
            round(job.processing_time_seconds, 1) if job.processing_time_seconds else None,
            job.segment_count,
            job.speaker_count,
            job.error_stage or "",
            (job.error_message or "")[:300],
            job.input_tokens or 0,
            job.output_tokens or 0,
        ])

    # Apply styling to all sheets
    for sheet in wb.worksheets:
        _style_headers(sheet)
        _auto_width(sheet)

    # Save to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"autoprotocol_stats_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
    return buf, filename
