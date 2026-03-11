"""
Business Domain Excel Generator.

Generates structured Excel from JSON result of business meeting analysis.
Universal generator for all Business meeting types.
"""
import logging
from typing import Optional, Any
from pathlib import Path

from openpyxl import Workbook

from backend.core.utils.excel_utils import create_sheet_with_table, create_info_sheet
from backend.domains.business.schemas import (
    BusinessMeetingType,
    NegotiationResult,
    ClientMeetingResult,
    StrategicPlanningResult,
    PresentationResult,
    WorkMeetingResult,
    BrainstormResult,
    LectureResult,
)

log = logging.getLogger(__name__)


# =============================================================================
# Per-type Excel generators
# =============================================================================

def _generate_negotiation_excel(result: NegotiationResult, wb: Workbook):
    info_fields = [("Цель встречи", result.meeting_goal)]
    if result.risk_level:
        info_fields.append(("Уровень рисков", result.risk_level))
    if result.internal_strategic_analysis:
        info_fields.append(("Внутренний стратегический анализ", result.internal_strategic_analysis))
    create_info_sheet(wb, "Инфо", info_fields)

    if result.parties:
        rows = []
        for party in result.parties:
            for rep in party.representatives:
                rows.append([party.party_name, rep])
            if not party.representatives:
                rows.append([party.party_name, "—"])
        create_sheet_with_table(wb, "Стороны", ["Сторона", "Представитель"], rows, [30, 30])

    if result.key_topics:
        rows = [[t.topic, t.positions, t.result] for t in result.key_topics]
        create_sheet_with_table(wb, "Темы", ["Тема", "Позиции сторон", "Итог"], rows, [30, 40, 20])

    if result.agreements:
        rows = [[a] for a in result.agreements]
        create_sheet_with_table(wb, "Договорённости", ["Договорённость"], rows, [80])

    if result.open_questions:
        rows = [[q] for q in result.open_questions]
        create_sheet_with_table(wb, "Открытые вопросы", ["Вопрос"], rows, [80])

    all_action_rows = []
    if result.action_items_for_us:
        for task in result.action_items_for_us:
            all_action_rows.append(["Наша сторона", task])
    if result.action_items_for_counterpart:
        for task in result.action_items_for_counterpart:
            all_action_rows.append(["Контрагент", task])
    if all_action_rows:
        create_sheet_with_table(wb, "Задачи", ["Сторона", "Задача"], all_action_rows, [20, 60])


def _generate_client_meeting_excel(result: ClientMeetingResult, wb: Workbook):
    info_fields = [("Цель встречи", result.meeting_goal)]
    if result.meeting_outcome:
        info_fields.append(("Итог встречи", result.meeting_outcome))
    if result.interest_level:
        info_fields.append(("Заинтересованность клиента", result.interest_level))
    if result.client_info:
        info_fields.append(("Компания клиента", result.client_info.company))
        if result.client_info.representatives:
            info_fields.append(("Представители", result.client_info.representatives))
    create_info_sheet(wb, "Инфо", info_fields)

    if result.client_needs:
        rows = [[n] for n in result.client_needs]
        create_sheet_with_table(wb, "Потребности", ["Потребность клиента"], rows, [80])

    if result.proposed_solutions:
        rows = [[s] for s in result.proposed_solutions]
        create_sheet_with_table(wb, "Решения", ["Предложенное решение"], rows, [80])

    if result.client_feedback:
        rows = [[f] for f in result.client_feedback]
        create_sheet_with_table(wb, "Обратная связь", ["Обратная связь"], rows, [80])

    if result.agreements:
        rows = [[a] for a in result.agreements]
        create_sheet_with_table(wb, "Договорённости", ["Договорённость"], rows, [80])

    if result.next_steps:
        rows = [[s.action, s.responsible or "—", s.deadline or "—"] for s in result.next_steps]
        create_sheet_with_table(wb, "Следующие шаги", ["Действие", "Ответственный", "Срок"], rows, [50, 25, 15])


def _generate_strategic_planning_excel(result: StrategicPlanningResult, wb: Workbook):
    create_info_sheet(wb, "Инфо", [
        ("Тема сессии", result.session_topic),
        ("Текущая ситуация", result.current_situation),
    ])

    if result.strategic_goals:
        rows = [[g] for g in result.strategic_goals]
        create_sheet_with_table(wb, "Цели", ["Стратегическая цель"], rows, [80])

    if result.initiatives:
        rows = [[i.name, i.priority, i.responsible or "—", i.timeline or "—"] for i in result.initiatives]
        create_sheet_with_table(wb, "Инициативы", ["Инициатива", "Приоритет", "Ответственный", "Сроки"], rows, [40, 15, 25, 20])

    if result.risks:
        rows = [[r] for r in result.risks]
        create_sheet_with_table(wb, "Риски", ["Риск / ограничение"], rows, [80])

    if result.kpis:
        rows = [[k.metric, k.target or "—"] for k in result.kpis]
        create_sheet_with_table(wb, "KPI", ["Метрика", "Целевое значение"], rows, [50, 30])

    if result.next_steps:
        rows = [[s.action, s.responsible or "—", s.deadline or "—"] for s in result.next_steps]
        create_sheet_with_table(wb, "Следующие шаги", ["Действие", "Ответственный", "Срок"], rows, [50, 25, 15])


def _generate_presentation_excel(result: PresentationResult, wb: Workbook):
    create_info_sheet(wb, "Инфо", [
        ("Название", result.title),
        ("Докладчик", result.presenter),
    ])

    if result.key_messages:
        rows = [[m] for m in result.key_messages]
        create_sheet_with_table(wb, "Ключевые тезисы", ["Тезис"], rows, [80])

    if result.conclusions:
        rows = [[c] for c in result.conclusions]
        create_sheet_with_table(wb, "Выводы", ["Вывод"], rows, [80])

    if result.audience_questions:
        rows = [[qa.question, qa.answer or "—"] for qa in result.audience_questions]
        create_sheet_with_table(wb, "Вопросы", ["Вопрос", "Ответ"], rows, [40, 50])

    if result.decisions:
        rows = [[d] for d in result.decisions]
        create_sheet_with_table(wb, "Решения", ["Решение"], rows, [80])

    if result.next_steps:
        rows = [[s.action, s.responsible or "—", s.deadline or "—"] for s in result.next_steps]
        create_sheet_with_table(wb, "Следующие шаги", ["Действие", "Ответственный", "Срок"], rows, [50, 25, 15])


def _generate_work_meeting_excel(result: WorkMeetingResult, wb: Workbook):
    create_info_sheet(wb, "Инфо", [
        ("Тема совещания", result.meeting_topic),
        ("Саммари", result.summary),
    ])

    if result.task_statuses:
        rows = [[ts.task, ts.responsible or "—", ts.status, ts.comment or "—"] for ts in result.task_statuses]
        create_sheet_with_table(wb, "Статусы задач", ["Задача", "Ответственный", "Статус", "Комментарий"], rows, [40, 20, 18, 30])

    if result.blockers:
        rows = [[b] for b in result.blockers]
        create_sheet_with_table(wb, "Блокеры", ["Блокер / проблема"], rows, [80])

    if result.decisions:
        rows = [[d] for d in result.decisions]
        create_sheet_with_table(wb, "Решения", ["Решение"], rows, [80])

    if result.action_items:
        rows = [[s.action, s.responsible or "—", s.deadline or "—"] for s in result.action_items]
        create_sheet_with_table(wb, "Поручения", ["Поручение", "Ответственный", "Срок"], rows, [50, 25, 15])


def _generate_brainstorm_excel(result: BrainstormResult, wb: Workbook):
    create_info_sheet(wb, "Инфо", [
        ("Тема сессии", result.session_topic),
        ("Проблема/задача", result.main_problem),
    ])

    if result.idea_clusters:
        rows = []
        for cluster in result.idea_clusters:
            for idea in cluster.ideas:
                rows.append([cluster.cluster_name, idea])
        if rows:
            create_sheet_with_table(wb, "Идеи по кластерам", ["Кластер", "Идея"], rows, [25, 60])

    if result.top_ideas:
        rows = [
            [idea.idea_description, idea.potential_impact or "—", idea.implementation_complexity or "—"]
            for idea in result.top_ideas
        ]
        create_sheet_with_table(wb, "Топ идеи", ["Идея", "Влияние", "Сложность"], rows, [50, 15, 15])

    if result.parked_ideas:
        rows = [[idea] for idea in result.parked_ideas]
        create_sheet_with_table(wb, "Парковка", ["Отложенная идея"], rows, [80])

    if result.next_steps:
        rows = [
            [step.action_item, step.responsible or "—", step.deadline or "—"]
            for step in result.next_steps
        ]
        create_sheet_with_table(wb, "Следующие шаги", ["Задача", "Ответственный", "Срок"], rows, [50, 25, 15])


def _generate_lecture_excel(result: LectureResult, wb: Workbook):
    create_info_sheet(wb, "Инфо", [
        ("Название", result.webinar_title),
    ])

    if result.presentation_part:
        rows = [
            [
                block.block_title,
                block.time_code or "—",
                block.key_idea or "—",
                "\n".join(block.theses) if block.theses else "—"
            ]
            for block in result.presentation_part
        ]
        create_sheet_with_table(wb, "Основная часть", ["Блок", "Тайм-код", "Ключевая мысль", "Тезисы"], rows, [25, 12, 40, 50])

    if result.qa_part:
        rows = [
            [
                qa.question_title,
                qa.time_code or "—",
                qa.key_answer_idea or "—",
                "\n".join(qa.answer_theses) if qa.answer_theses else "—"
            ]
            for qa in result.qa_part
        ]
        create_sheet_with_table(wb, "Q&A", ["Вопрос", "Тайм-код", "Ключевая мысль ответа", "Тезисы"], rows, [30, 12, 40, 50])

    if result.final_summary:
        rows = [[item] for item in result.final_summary]
        create_sheet_with_table(wb, "Выводы", ["Вывод"], rows, [80])


# =============================================================================
# Main dispatcher
# =============================================================================

def generate_business_excel(
    meeting_type: BusinessMeetingType,
    result: Any,
    output_path: Path,
    meeting_date: Optional[str] = None
) -> Path:
    """
    Generate Excel report for any Business meeting type.

    Args:
        meeting_type: Type of Business meeting
        result: Parsed result object
        output_path: Path to save the Excel file
        meeting_date: Optional meeting date

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    generators = {
        BusinessMeetingType.NEGOTIATION: _generate_negotiation_excel,
        BusinessMeetingType.CLIENT_MEETING: _generate_client_meeting_excel,
        BusinessMeetingType.STRATEGIC_PLANNING: _generate_strategic_planning_excel,
        BusinessMeetingType.PRESENTATION: _generate_presentation_excel,
        BusinessMeetingType.WORK_MEETING: _generate_work_meeting_excel,
        BusinessMeetingType.BRAINSTORM: _generate_brainstorm_excel,
        BusinessMeetingType.LECTURE: _generate_lecture_excel,
    }

    generator = generators.get(meeting_type)
    if not generator:
        raise ValueError(f"Unknown Business meeting type: {meeting_type}")

    generator(result, wb)

    # Fallback if no sheets were created
    if not wb.sheetnames:
        create_info_sheet(wb, "Инфо", [
            ("Тип встречи", meeting_type.value),
            ("Дата", meeting_date or "—"),
            ("Данные", "Нет структурированных данных"),
        ])

    wb.save(str(output_path))
    log.info(f"Business Excel saved: {output_path}")
    return output_path
