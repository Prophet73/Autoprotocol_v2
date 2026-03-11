"""
CEO Domain Excel Generator.

Generates structured Excel from JSON result of CEO NOTECH meeting analysis.
"""
import logging
from typing import Optional, Any
from pathlib import Path

from openpyxl import Workbook

from backend.core.utils.excel_utils import create_sheet_with_table, create_info_sheet
from backend.domains.ceo.schemas import CEOMeetingType, NotechResult, NotechQuestion

log = logging.getLogger(__name__)


def _generate_notech_excel(result: NotechResult, wb: Workbook):
    """Generate Excel sheets for NOTECH meeting report."""
    # 1. Инфо
    info_fields = [
        ("Тема совещания", result.meeting_topic),
        ("Саммари", result.summary),
    ]
    if result.attendees:
        info_fields.append(("Участники", result.attendees))
    create_info_sheet(wb, "Инфо", info_fields)

    # 2. Вопросы повестки — основной контент
    if result.questions:
        rows = []
        for i, q in enumerate(result.questions, 1):
            rows.append([
                f"Вопрос №{i}: {q.title}",
                q.description,
                q.decision or "—",
                "\n".join(q.value_points) if q.value_points else "—",
                "\n".join(q.discussion_details) if q.discussion_details else "—",
                "\n".join(q.risks) if q.risks else "—",
            ])
        create_sheet_with_table(
            wb, "Вопросы повестки",
            ["Вопрос", "Суть проблемы", "Решение", "Ценность", "Детали обсуждения", "Риски"],
            rows, [35, 40, 35, 30, 40, 30],
        )

    # 3. Принятые решения / задачи
    if result.action_items:
        rows = [[item] for item in result.action_items]
        create_sheet_with_table(
            wb, "Принятые решения",
            ["Задача / поручение"], rows, [80],
        )


# =============================================================================
# Main dispatcher
# =============================================================================

def generate_ceo_excel(
    meeting_type: CEOMeetingType,
    result: Any,
    output_path: Path,
    meeting_date: Optional[str] = None
) -> Path:
    """
    Generate Excel report for CEO meeting type.

    Args:
        meeting_type: Type of CEO meeting
        result: Parsed result object
        output_path: Path to save the Excel file
        meeting_date: Optional meeting date

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    generators = {
        CEOMeetingType.NOTECH: _generate_notech_excel,
    }

    generator = generators.get(meeting_type)
    if not generator:
        raise ValueError(f"Unknown CEO meeting type: {meeting_type}")

    generator(result, wb)

    # Fallback if no sheets were created
    if not wb.sheetnames:
        create_info_sheet(wb, "Инфо", [
            ("Тип встречи", meeting_type.value),
            ("Дата", meeting_date or "—"),
            ("Данные", "Нет структурированных данных"),
        ])

    wb.save(str(output_path))
    log.info(f"CEO Excel saved: {output_path}")
    return output_path
