"""
Tasks generator - creates tasks.xlsx from BasicReport.
Receives pre-generated BasicReport (from shared LLM call) and formats as Excel.
"""

import logging
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.core.transcription.models import TranscriptionResult
from backend.domains.construction.schemas import BasicReport, TaskCategory, TaskPriority, TaskConfidence


logger = logging.getLogger(__name__)


def generate_tasks(
    result: TranscriptionResult,
    output_dir: Path,
    basic_report: BasicReport,
    filename: str = None,
    participants: list = None,
    meeting_date: str = None,
) -> Path:
    """
    Generate tasks.xlsx from BasicReport.

    Args:
        result: TranscriptionResult from pipeline (for metadata)
        output_dir: Directory to save the file
        basic_report: Pre-generated BasicReport from shared LLM call
        filename: Optional custom filename
        participants: Optional list of participants
        meeting_date: Optional meeting date string (YYYY-MM-DD)

    Returns:
        Path to generated file
    """

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row with meeting date
    title_text = "Извлечённые задачи"
    if meeting_date:
        try:
            dt = datetime.strptime(meeting_date, "%Y-%m-%d")
            title_text = f"Задачи совещания от {dt.strftime('%d.%m.%Y')}"
        except ValueError:
            title_text = f"Задачи совещания от {meeting_date}"

    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # Headers (row 2 now)
    headers = ["№", "Уверенность", "Приоритет", "Категория", "Задача", "Ответственный", "Срок", "Примечания", "Тайм-код", "Источник"]
    col_widths = [5, 14, 12, 22, 50, 20, 15, 25, 15, 40]

    # Clean style - no rainbow colors, just alternating rows for readability
    alt_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Light gray

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sort tasks by category (alphabetically), then by priority within category
    priority_order = {"high": 0, "medium": 1, "low": 2}

    def get_sort_key(task):
        # Get category string for sorting
        cat_val = task.category.value if isinstance(task.category, TaskCategory) else str(task.category)
        # Get priority order for sorting within category
        pri_val = task.priority.value if isinstance(task.priority, TaskPriority) else str(task.priority)
        pri_order = priority_order.get(pri_val, 1)
        return (cat_val, pri_order)

    sorted_tasks = sorted(basic_report.tasks, key=get_sort_key)

    # Data rows (start from row 3, after title + header)
    for row_num, task in enumerate(sorted_tasks, 3):
        # Priority label
        priority_val = task.priority.value if isinstance(task.priority, TaskPriority) else str(task.priority)
        priority_labels = {"high": "Высокий", "medium": "Средний", "low": "Низкий"}
        priority_label = priority_labels.get(priority_val, priority_val)

        # Confidence label
        confidence_val = task.confidence.value if isinstance(task.confidence, TaskConfidence) else str(getattr(task, 'confidence', 'high'))
        confidence_labels = {"high": "Явная", "medium": "Из контекста"}
        confidence_label = confidence_labels.get(confidence_val, confidence_val)

        # Time codes as string
        time_codes_str = ", ".join(task.time_codes) if task.time_codes else ""

        row_data = [
            row_num - 2,
            confidence_label,
            priority_label,
            task.category.value if isinstance(task.category, TaskCategory) else str(task.category),
            task.description,
            task.responsible or "",
            task.deadline or "",
            task.notes or "",
            time_codes_str,
            task.evidence or "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Alternating row colors for readability (no rainbow)
            if row_num % 2 == 0:
                cell.fill = alt_row_fill

    # Add metadata sheet
    ws_meta = wb.create_sheet("Метаданные")
    metadata = [
        ("Исходный файл", result.metadata.source_file),
        ("Длительность", result.metadata.duration_formatted),
        ("Дата обработки", datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")),
        ("Тип совещания", basic_report.meeting_type),
        ("Краткое содержание", basic_report.meeting_summary),
        ("Экспертный анализ", basic_report.expert_analysis),
    ]

    for row, (label, value) in enumerate(metadata, 1):
        ws_meta.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_meta.cell(row=row, column=2, value=value)

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 80

    # Add participants sheet if provided
    if participants:
        ws_part = wb.create_sheet("Участники")

        # Role labels mapping
        role_labels = {
            "GENERAL": "Генподрядчик",
            "CUSTOMER": "Заказчик",
            "ЗАКАЗЧИК": "Заказчик",
            "TECHNICAL_CUSTOMER": "Технический заказчик",
            "ТЕХНИЧЕСКИЙ ЗАКАЗЧИК": "Технический заказчик",
            "DESIGNER": "Проектировщик",
            "SUBCONTRACTOR": "Субподрядчик",
            "SUPPLIER": "Поставщик",
            "INVESTOR": "Инвестор",
        }

        # Headers
        part_headers = ["Организация", "Роль", "ФИО", "Должность"]
        for col, header in enumerate(part_headers, 1):
            cell = ws_part.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        row_num = 2
        for org in participants:
            org_name = org.get("organization", "")
            role_raw = org.get("role", "")
            role = role_labels.get(role_raw.upper(), role_raw)

            for person in org.get("people", []):
                ws_part.cell(row=row_num, column=1, value=org_name).border = thin_border
                ws_part.cell(row=row_num, column=2, value=role).border = thin_border
                # Handle both formats: string "Name (Position)" or dict {"name": ..., "position": ...}
                if isinstance(person, str):
                    # Parse "Иванов Иван Петрови (Директор)" format
                    if "(" in person and person.endswith(")"):
                        name_part = person[:person.rfind("(")].strip()
                        position_part = person[person.rfind("(")+1:-1].strip()
                    else:
                        name_part = person
                        position_part = ""
                    ws_part.cell(row=row_num, column=3, value=name_part).border = thin_border
                    ws_part.cell(row=row_num, column=4, value=position_part).border = thin_border
                else:
                    ws_part.cell(row=row_num, column=3, value=person.get("name", "")).border = thin_border
                    ws_part.cell(row=row_num, column=4, value=person.get("position", "")).border = thin_border
                row_num += 1

        # Column widths
        ws_part.column_dimensions["A"].width = 30
        ws_part.column_dimensions["B"].width = 20
        ws_part.column_dimensions["C"].width = 30
        ws_part.column_dimensions["D"].width = 25

    # Save workbook
    output_dir.mkdir(parents=True, exist_ok=True)

    if filename is None:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"tasks_{timestamp}.xlsx"

    output_path = output_dir / filename
    wb.save(str(output_path))

    return output_path


def regenerate_tasks_xlsx(
    basic_report: BasicReport,
    output_path: Path,
    source_file: str = "N/A",
    duration: str = "N/A",
    participants: list = None,
    meeting_date: str = None,
) -> Path:
    """
    Regenerate tasks.xlsx from edited BasicReport JSON.

    Used when manager edits basic_report_json in DB and needs to regenerate the XLSX.

    Args:
        basic_report: BasicReport object (can be created from edited JSON)
        output_path: Full path to output XLSX file
        source_file: Original source filename
        duration: Duration string
        participants: Optional list of participants
        meeting_date: Optional meeting date string (YYYY-MM-DD)

    Returns:
        Path to generated XLSX file
    """
    from backend.core.transcription.models import TranscriptionResult, TranscriptionMetadata

    # Create minimal TranscriptionResult for metadata
    metadata = TranscriptionMetadata(
        source_file=source_file,
        duration_seconds=0.0,
    )
    # Override duration_formatted
    metadata._duration_formatted = duration

    result = TranscriptionResult(
        metadata=metadata,
        segments=[],
        speakers={},
    )

    # Call existing generate_tasks function
    output_dir = output_path.parent
    filename = output_path.name

    return generate_tasks(
        result=result,
        output_dir=output_dir,
        basic_report=basic_report,
        filename=filename,
        participants=participants,
        meeting_date=meeting_date,
    )
