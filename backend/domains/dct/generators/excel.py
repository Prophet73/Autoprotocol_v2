"""
DCT Domain Excel Generator.

Генерирует структурированный Excel из JSON-результата анализа встречи.
Универсальный генератор для всех типов встреч DCT.
"""
import logging
from typing import Optional, Any
from pathlib import Path

from openpyxl import Workbook

from backend.core.utils.excel_utils import create_sheet_with_table, create_info_sheet
from backend.domains.dct.schemas import (
    DCTMeetingType,
    BrainstormResult,
    ProductionMeetingResult,
    NegotiationResult,
    LectureResult,
)

log = logging.getLogger(__name__)


def _generate_brainstorm_excel(result: BrainstormResult, wb: Workbook):
    """Генерация листов для мозгового штурма."""
    create_info_sheet(wb, "Инфо", [
        ("Тема сессии", result.session_topic),
        ("Проблема/задача", result.main_problem),
    ])

    if result.idea_clusters:
        rows = []
        for cluster in result.idea_clusters:
            for idea in cluster.ideas:
                rows.append([cluster.cluster_name, idea])
        if rows:
            create_sheet_with_table(
                wb, "Идеи по кластерам",
                ["Кластер", "Идея"],
                rows,
                [25, 60]
            )

    if result.top_ideas:
        rows = [
            [idea.idea_description, idea.potential_impact or "—", idea.implementation_complexity or "—"]
            for idea in result.top_ideas
        ]
        create_sheet_with_table(
            wb, "Топ идеи",
            ["Идея", "Влияние", "Сложность"],
            rows,
            [50, 15, 15]
        )

    if result.parked_ideas:
        rows = [[idea] for idea in result.parked_ideas]
        create_sheet_with_table(
            wb, "Парковка",
            ["Отложенная идея"],
            rows,
            [80]
        )

    if result.next_steps:
        rows = [
            [step.action_item, step.responsible or "—", step.deadline or "—"]
            for step in result.next_steps
        ]
        create_sheet_with_table(
            wb, "Следующие шаги",
            ["Задача", "Ответственный", "Срок"],
            rows,
            [50, 25, 15]
        )


def _generate_production_excel(result: ProductionMeetingResult, wb: Workbook):
    """Генерация листов для производственного совещания."""
    info_fields = [("Объект", result.object_name)]
    if result.summary:
        info_fields.append(("Саммари", result.summary))
    info_fields.append(("Участники", result.attendees))
    create_info_sheet(wb, "Инфо", info_fields)

    if result.past_tasks_control:
        rows = [
            [task.task_description, task.status, task.comment or "—"]
            for task in result.past_tasks_control
        ]
        create_sheet_with_table(
            wb, "Контроль задач",
            ["Задача", "Статус", "Комментарий"],
            rows,
            [50, 15, 30]
        )

    if result.work_progress_analysis:
        rows = [
            [item.work_block_name, item.status_summary]
            for item in result.work_progress_analysis
        ]
        create_sheet_with_table(
            wb, "Ход работ",
            ["Блок работ", "Статус"],
            rows,
            [30, 60]
        )

    if result.resources_and_supply:
        res = result.resources_and_supply
        create_info_sheet(wb, "Ресурсы", [
            ("Людские ресурсы", res.manpower),
            ("Техника", res.machinery),
            ("Материалы", res.materials),
        ])

    if result.safety_and_labor_protection:
        rows = [[item] for item in result.safety_and_labor_protection]
        create_sheet_with_table(
            wb, "ОТ и ТБ",
            ["Вопрос"],
            rows,
            [80]
        )

    if result.new_tasks:
        rows = [
            [task.task_description, task.responsible, task.deadline or "—"]
            for task in result.new_tasks
        ]
        create_sheet_with_table(
            wb, "Новые задачи",
            ["Задача", "Ответственный", "Срок"],
            rows,
            [50, 25, 15]
        )


def _generate_negotiation_excel(result: NegotiationResult, wb: Workbook):
    """Генерация листов для переговоров."""
    create_info_sheet(wb, "Инфо", [
        ("Цель встречи", result.meeting_goal),
        ("Контрагент", result.counterpart_company),
    ])

    if result.topics_discussed:
        rows = []
        for topic in result.topics_discussed:
            rows.append([
                topic.topic_title,
                topic.proposal_summary or "—",
                "\n".join(topic.value_for_company) if topic.value_for_company else "—",
                "\n".join(topic.risks_and_objections) if topic.risks_and_objections else "—",
                "\n".join(topic.terms_and_cost) if topic.terms_and_cost else "—",
            ])
        create_sheet_with_table(
            wb, "Темы переговоров",
            ["Тема", "Суть предложения", "Ценность для нас", "Риски", "Условия"],
            rows,
            [25, 35, 30, 30, 25]
        )

    if result.action_items:
        rows = []
        if result.action_items.for_us:
            for task in result.action_items.for_us:
                rows.append(["Наша сторона", task])
        if result.action_items.for_counterpart:
            for task in result.action_items.for_counterpart:
                rows.append(["Контрагент", task])
        if rows:
            create_sheet_with_table(
                wb, "Задачи",
                ["Сторона", "Задача"],
                rows,
                [20, 60]
            )

    if result.internal_strategic_analysis:
        create_info_sheet(wb, "Стратегический анализ", [
            ("Анализ", result.internal_strategic_analysis),
        ])


def _generate_lecture_excel(result: LectureResult, wb: Workbook):
    """Генерация листов для лекции/вебинара."""
    create_info_sheet(wb, "Инфо", [
        ("Название", result.webinar_title),
    ])

    if result.presentation_part:
        rows = [
            [
                block.block_title,
                block.time_code or "—",
                block.key_idea or "—",
                "\n".join(block.theses) if block.theses else "—"
            ]
            for block in result.presentation_part
        ]
        create_sheet_with_table(
            wb, "Основная часть",
            ["Блок", "Тайм-код", "Ключевая мысль", "Тезисы"],
            rows,
            [25, 12, 40, 50]
        )

    if result.qa_part:
        rows = [
            [
                qa.question_title,
                qa.time_code or "—",
                qa.key_answer_idea or "—",
                "\n".join(qa.answer_theses) if qa.answer_theses else "—"
            ]
            for qa in result.qa_part
        ]
        create_sheet_with_table(
            wb, "Q&A",
            ["Вопрос", "Тайм-код", "Ключевая мысль ответа", "Тезисы"],
            rows,
            [30, 12, 40, 50]
        )

    if result.final_summary:
        rows = [[item] for item in result.final_summary]
        create_sheet_with_table(
            wb, "Выводы",
            ["Вывод"],
            rows,
            [80]
        )


def generate_dct_excel(
    meeting_type: DCTMeetingType,
    result: Any,
    output_path: Path,
    meeting_date: Optional[str] = None
) -> Path:
    """
    Генерация Excel-отчёта для любого типа встречи DCT.

    Args:
        meeting_type: Тип встречи DCT
        result: Объект результата (BrainstormResult, ProductionMeetingResult, etc.)
        output_path: Путь для сохранения файла
        meeting_date: Опциональная дата встречи

    Returns:
        Путь к созданному Excel файлу
    """
    wb = Workbook()
    wb.remove(wb.active)

    if meeting_type == DCTMeetingType.BRAINSTORM:
        _generate_brainstorm_excel(result, wb)
    elif meeting_type == DCTMeetingType.PRODUCTION:
        _generate_production_excel(result, wb)
    elif meeting_type == DCTMeetingType.NEGOTIATION:
        _generate_negotiation_excel(result, wb)
    elif meeting_type == DCTMeetingType.LECTURE:
        _generate_lecture_excel(result, wb)
    else:
        raise ValueError(f"Unknown DCT meeting type: {meeting_type}")

    if not wb.sheetnames:
        create_info_sheet(wb, "Инфо", [
            ("Тип встречи", meeting_type.value),
            ("Дата", meeting_date or "—"),
            ("Данные", "Нет структурированных данных"),
        ])

    wb.save(str(output_path))
    log.info(f"DCT Excel saved: {output_path}")
    return output_path
