"""
FTA Domain Excel Generator.

Generates structured Excel from JSON result of FTA audit analysis.
"""
import logging
from typing import Optional, Any
from pathlib import Path

from openpyxl import Workbook

from backend.core.utils.excel_utils import create_sheet_with_table, create_info_sheet
from backend.domains.fta.schemas import FTAMeetingType, AuditResult

log = logging.getLogger(__name__)


def _generate_audit_excel(result: AuditResult, wb: Workbook):
    """Generate Excel sheets for audit report."""
    info_fields = [
        ("Предмет аудита", result.audit_subject),
    ]
    if result.audit_scope:
        info_fields.append(("Охват проверки", result.audit_scope))
    if result.overall_rating:
        info_fields.append(("Общая оценка", result.overall_rating))
    if result.participants:
        info_fields.append(("Участники", result.participants))
    create_info_sheet(wb, "Инфо", info_fields)

    if result.findings:
        rows = [
            [f.finding, f.severity, f.area or "—", f.recommendation or "—"]
            for f in result.findings
        ]
        create_sheet_with_table(
            wb, "Замечания",
            ["Замечание", "Критичность", "Область", "Рекомендация"],
            rows, [50, 18, 25, 40],
        )

    if result.positive_observations:
        rows = [[obs] for obs in result.positive_observations]
        create_sheet_with_table(
            wb, "Положительные наблюдения",
            ["Наблюдение"], rows, [80],
        )

    if result.risks:
        rows = [[r] for r in result.risks]
        create_sheet_with_table(wb, "Риски", ["Риск"], rows, [80])

    if result.corrective_actions:
        rows = [
            [a.action, a.responsible or "—", a.deadline or "—"]
            for a in result.corrective_actions
        ]
        create_sheet_with_table(
            wb, "Корректирующие мероприятия",
            ["Мероприятие", "Ответственный", "Срок"],
            rows, [50, 25, 15],
        )

    if result.conclusions:
        rows = [[c] for c in result.conclusions]
        create_sheet_with_table(wb, "Выводы", ["Вывод"], rows, [80])


# =============================================================================
# Main dispatcher
# =============================================================================

def generate_fta_excel(
    meeting_type: FTAMeetingType,
    result: Any,
    output_path: Path,
    meeting_date: Optional[str] = None
) -> Path:
    """
    Generate Excel report for FTA meeting type.

    Args:
        meeting_type: Type of FTA meeting
        result: Parsed result object
        output_path: Path to save the Excel file
        meeting_date: Optional meeting date

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    generators = {
        FTAMeetingType.AUDIT: _generate_audit_excel,
    }

    generator = generators.get(meeting_type)
    if not generator:
        raise ValueError(f"Unknown FTA meeting type: {meeting_type}")

    generator(result, wb)

    # Fallback if no sheets were created
    if not wb.sheetnames:
        create_info_sheet(wb, "Инфо", [
            ("Тип встречи", meeting_type.value),
            ("Дата", meeting_date or "—"),
            ("Данные", "Нет структурированных данных"),
        ])

    wb.save(str(output_path))
    log.info(f"FTA Excel saved: {output_path}")
    return output_path
