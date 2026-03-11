"""
Shared Excel utilities for domain report generators.

Contains common styles and helper functions used across all domain Excel generators.
"""
from typing import Optional, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_sheet_with_table(
    wb: Workbook,
    sheet_name: str,
    headers: List[str],
    rows: List[List[Any]],
    col_widths: Optional[List[int]] = None
):
    """Create a sheet with a styled table."""
    ws = wb.create_sheet(sheet_name)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    if col_widths:
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

    return ws


def create_info_sheet(
    wb: Workbook,
    sheet_name: str,
    fields: List[tuple],
):
    """Create a sheet with key-value info."""
    ws = wb.create_sheet(sheet_name)

    for row_idx, (label, value) in enumerate(fields, 1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_label.alignment = Alignment(vertical="top")

        if isinstance(value, list):
            value = "\n".join(str(v) for v in value) if value else "—"
        cell_value = ws.cell(row=row_idx, column=2, value=value or "—")
        cell_value.alignment = CELL_ALIGNMENT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    return ws
