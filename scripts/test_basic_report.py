#!/usr/bin/env python3
"""
Test script for basic_report prompt.
Reads a transcript docx and generates tasks.xlsx using updated prompt.

Usage:
    python scripts/test_basic_report.py <path_to_transcript.docx>
"""
import os
import sys
import json
import yaml
from pathlib import Path
from datetime import datetime

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Load .env file
from dotenv import load_dotenv
env_path = project_root / ".env"
if env_path.exists():
    load_dotenv(env_path)
    print(f"Loaded .env from {env_path}")

from docx import Document
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def read_docx_transcript(docx_path: str) -> str:
    """Read transcript text from docx file."""
    doc = Document(docx_path)
    paragraphs = []
    for para in doc.paragraphs:
        if para.text.strip():
            paragraphs.append(para.text)
    return "\n".join(paragraphs)


def load_prompts() -> dict:
    """Load prompts from prompts.yaml."""
    prompts_path = Path(__file__).parent.parent / "backend" / "config" / "prompts.yaml"
    with open(prompts_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_basic_report_schema() -> dict:
    """Get JSON schema for BasicReport."""
    return {
        "type": "object",
        "properties": {
            "meeting_type": {
                "type": "string",
                "enum": ["production", "quality", "finance", "design", "coordination"]
            },
            "meeting_summary": {"type": "string"},
            "expert_analysis": {"type": "string"},
            "tasks": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "category": {
                            "type": "string",
                            "enum": ["ИРД", "Проектирование и РД", "СМР", "Инженерные системы",
                                     "ОТ и ТБ", "Финансы", "Взаимодействие", "Организация"]
                        },
                        "description": {"type": "string"},
                        "responsible": {"type": "string"},
                        "deadline": {"type": "string"},
                        "notes": {"type": "string"},
                        "priority": {"type": "string", "enum": ["high", "medium", "low"]},
                        "confidence": {"type": "string", "enum": ["high", "medium"]},
                        "time_codes": {"type": "array", "items": {"type": "string"}},
                        "evidence": {"type": "string"}
                    },
                    "required": ["category", "description", "priority", "confidence"]
                }
            }
        },
        "required": ["meeting_type", "meeting_summary", "expert_analysis", "tasks"]
    }


def extract_tasks_via_llm(transcript: str, meeting_date: str) -> dict:
    """Call Gemini to extract tasks from transcript."""

    # Check API key
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GOOGLE_API_KEY or GEMINI_API_KEY environment variable not set")

    # Initialize client
    client = genai.Client(api_key=api_key)

    # Load prompts
    prompts = load_prompts()
    construction_prompts = prompts.get("domains", {}).get("construction", {}).get("basic_report", {})

    system_prompt = construction_prompts.get("system", "")
    user_prompt_template = construction_prompts.get("user", "")

    if not user_prompt_template:
        raise ValueError("basic_report user prompt not found in prompts.yaml")

    # Format prompt
    user_prompt = user_prompt_template.format(
        transcript=transcript[:20000],  # Limit transcript size
        meeting_date=meeting_date
    )

    print(f"\n{'='*60}")
    print("SYSTEM PROMPT:")
    print('='*60)
    print(system_prompt[:500] + "..." if len(system_prompt) > 500 else system_prompt)

    print(f"\n{'='*60}")
    print("USER PROMPT (first 1000 chars):")
    print('='*60)
    print(user_prompt[:1000] + "...")

    # Call Gemini
    print(f"\n{'='*60}")
    print("Calling Gemini API...")
    print('='*60)

    # Use gemini-2.5-pro for quality report generation
    models_to_try = [
        "gemini-2.5-pro",
    ]

    response = None
    last_error = None

    for model in models_to_try:
        print(f"Trying model: {model}")
        try:
            import time
            time.sleep(2)  # Rate limit delay
            response = client.models.generate_content(
                model=model,
                contents=[system_prompt, user_prompt] if system_prompt else user_prompt,
                config={
                    "response_mime_type": "application/json",
                    "response_schema": get_basic_report_schema(),
                },
            )
            print(f"Success with model: {model}")
            break
        except Exception as e:
            last_error = e
            print(f"Failed with {model}: {e}")
            continue

    if response is None:
        raise last_error or Exception("All models failed")

    # Parse response
    result = json.loads(response.text)
    return result


def generate_excel(report: dict, output_path: Path, source_file: str):
    """Generate Excel file from BasicReport."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["№", "Уверенность", "Приоритет", "Категория", "Задача", "Ответственный", "Срок", "Примечания", "Тайм-код", "Источник"]
    col_widths = [5, 14, 12, 22, 50, 20, 15, 25, 15, 40]

    # Priority colors
    priority_fills = {
        "high": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "low": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    }

    # Confidence fills
    confidence_fills = {
        "high": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "medium": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    }

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    tasks = report.get("tasks", [])
    for row_num, task in enumerate(tasks, 2):
        priority_val = task.get("priority", "medium")
        priority_labels = {"high": "Высокий", "medium": "Средний", "low": "Низкий"}
        priority_label = priority_labels.get(priority_val, priority_val)

        confidence_val = task.get("confidence", "high")
        confidence_labels = {"high": "Явная", "medium": "Из контекста"}
        confidence_label = confidence_labels.get(confidence_val, confidence_val)

        time_codes = task.get("time_codes", [])
        time_codes_str = ", ".join(time_codes) if time_codes else ""

        row_data = [
            row_num - 1,
            confidence_label,
            priority_label,
            task.get("category", ""),
            task.get("description", ""),
            task.get("responsible") or "",
            task.get("deadline") or "",
            task.get("notes") or "",
            time_codes_str,
            task.get("evidence") or "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if priority_val in priority_fills:
                cell.fill = priority_fills[priority_val]

            if col == 2 and confidence_val in confidence_fills:
                cell.fill = confidence_fills[confidence_val]

    # Metadata sheet
    ws_meta = wb.create_sheet("Метаданные")
    metadata = [
        ("Исходный файл", source_file),
        ("Дата обработки", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Тип совещания", report.get("meeting_type", "")),
        ("Краткое содержание", report.get("meeting_summary", "")),
        ("Экспертный анализ", report.get("expert_analysis", "")),
        ("Всего задач", len(tasks)),
        ("Явных задач (high)", len([t for t in tasks if t.get("confidence") == "high"])),
        ("Из контекста (medium)", len([t for t in tasks if t.get("confidence") == "medium"])),
    ]

    for row, (label, value) in enumerate(metadata, 1):
        ws_meta.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_meta.cell(row=row, column=2, value=value)

    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 80

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/test_basic_report.py <path_to_transcript.docx>")
        sys.exit(1)

    docx_path = sys.argv[1]

    if not os.path.exists(docx_path):
        print(f"File not found: {docx_path}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"Reading transcript: {docx_path}")
    print('='*60)

    # Read transcript
    transcript = read_docx_transcript(docx_path)
    print(f"Transcript length: {len(transcript)} chars")
    print(f"First 500 chars:\n{transcript[:500]}...")

    # Extract meeting date from filename or use today
    meeting_date = datetime.now().strftime("%d.%m.%Y")

    # Extract tasks via LLM
    report = extract_tasks_via_llm(transcript, meeting_date)

    print(f"\n{'='*60}")
    print("RESULTS:")
    print('='*60)
    print(f"Meeting type: {report.get('meeting_type')}")
    print(f"Summary: {report.get('meeting_summary')}")
    print(f"Expert analysis: {report.get('expert_analysis')}")
    print(f"\nTotal tasks: {len(report.get('tasks', []))}")

    tasks = report.get("tasks", [])
    high_conf = [t for t in tasks if t.get("confidence") == "high"]
    medium_conf = [t for t in tasks if t.get("confidence") == "medium"]

    print(f"  - Явных (high confidence): {len(high_conf)}")
    print(f"  - Из контекста (medium confidence): {len(medium_conf)}")

    # Print task breakdown
    print(f"\n{'='*60}")
    print("TASKS BY CATEGORY:")
    print('='*60)

    categories = {}
    for task in tasks:
        cat = task.get("category", "Unknown")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(task)

    for cat, cat_tasks in sorted(categories.items()):
        print(f"\n{cat} ({len(cat_tasks)}):")
        for t in cat_tasks:
            conf_marker = "✓" if t.get("confidence") == "high" else "?"
            print(f"  [{conf_marker}] {t.get('description', '')[:80]}...")

    # Generate Excel
    source_name = Path(docx_path).stem
    output_path = Path(docx_path).parent / f"{source_name}_tasks.xlsx"

    generate_excel(report, output_path, Path(docx_path).name)

    print(f"\n{'='*60}")
    print(f"Excel saved to: {output_path}")
    print('='*60)

    # Also save JSON
    json_path = Path(docx_path).parent / f"{source_name}_report.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"JSON saved to: {json_path}")


if __name__ == "__main__":
    main()
